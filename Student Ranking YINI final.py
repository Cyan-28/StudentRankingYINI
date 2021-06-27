# Importing
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook("data.xlsx")
sheet = wb.active

# Arrays
studentID = []
commonData = [[], [], [], []] # Data needed for progress score and normal average
progSpecData = [[], [], [], []]  # Progress Specifc Data

# Functions and Classes


def getData(currentArray, notNeededArray, startColNum):
    counter = 0
    for x in range(0, sheet.max_column, 3):
        character = get_column_letter(x + startColNum)
        for col in range(2, sheet.max_row+1):
            if counter < 3:
                currentArray[counter].append(sheet[character+str(col)].value)
            else:
                notNeededArray[counter].append(sheet[character+str(col)].value)
        counter += 1


class Average():

    def __init__(self, array, student):
        self.array = array
        self.student = student

    def Calculate(self):
        for x in range(0, len(self.array[0])):
            average = 0
            for n in range(0,len(self.array)-1):
                average += self.array[n][x]
                if n == 2:
                    average = round(average/3,2)
                    #assert self.array[3][x] == average # Works but commented in favor of visual
                    if average == self.array[3][x]:
                        print("Student",self.student[x],"Average is correct")
                    else:
                        print("Student",self.student[x],"Average is incorrect")


class Progress():

    def __init__(self, progArray, commArray, studentArray):

        self.progArray = progArray
        self.commArray = commArray
        self.student = studentArray

    def Calculate(self):
        for x in range(0, len(self.progArray[0])):
            progAv = 0
            for n in range(0,len(self.progArray)-1):
                progAv += ( ((self.commArray[n][x]) - (self.progArray[n][x])) / 10 )
                if n == 2:
                    progAv = round(progAv/3,2)
                    #assert self.progArray[3][x] == progAv # Works but commented in favor of visual
                    if progAv == self.progArray[3][x]:
                        print("Student",self.student[x],"Prog Average is correct")
                    else:
                        print("Student",self.student[x],"Prog Average is incorrect")


# Reading Excel File
# StudentID
for col in range(2, sheet.max_row+1):
    char = "A"
    studentID.append(sheet[char + str(col)].value)

# CommonData and Prog Data
getData(commonData, progSpecData, 3)
getData(progSpecData, commonData, 2)
print(commonData)
print(progSpecData)


testing = Average(commonData, studentID)
testing.Calculate()


testing2 = Progress(progSpecData, commonData, studentID)
testing2.Calculate()