class AvScore:

    def __init__(self, maths, english, science):

        self.maths = maths
        self.english = english
        self.science = science

    def CalculateAverage(self):

        done = (self.maths + self.science + self.english)/3
        print("Average:", done)

class ProgScore:

    def __init__(self, maths, mMaths, english, mEnglish, science, mScience):

        self.maths = maths
        self.mMaths = mMaths
        self.english = english
        self.mEnglish = mEnglish
        self.science = science
        self.mScience = mScience

    def ProgressScore(self):

        progMaths = (self.maths-self.mMaths)/10
        progEnglish = (self.english-self.mEnglish)/10
        progScience = (self.science-self.mScience)/10
        progAv = (progMaths+progEnglish+progScience)/3

        print("Progress Maths:", str(round(progMaths, 2)), "\nProgress English:", str(round(progEnglish, 2)),
              "\nProgress Science:", str(round(progScience, 2)), "\nProgress Average:", str(round(progAv, 2)))


#p = ProgScore(88, 99, 96, 92, 54, 79)
#p.ProgressScore()


# XL
from openpyxl import load_workbook
wb = load_workbook("data.xlsx")
sheet = wb.active

print(f"Max row in the active sheet is {sheet.max_row}")
print(f"Max column in the active sheet is {sheet.max_column}")

table=[]

for x in range(2,sheet.max_row+1):
    for data in sheet[x]:
        table.append[x][data.value]